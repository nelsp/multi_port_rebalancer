#!/usr/bin/env python3
"""
Multi-Account Crypto Portfolio Rebalancer
==========================================
Reads a single market data CSV and multiple account JSON files,
computes per-token analytics (volatility, 30d averages, covariance),
then generates one Excel rebalancing report per account.

Usage:
    python rebalancer.py
    (prompts for data file path, then processes all accounts in accounts/)
"""

import pandas as pd
import numpy as np
import json
import os
import sys
import glob
from datetime import datetime
from scipy.optimize import minimize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ═══════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
ACCOUNTS_DIR = os.path.join(SCRIPT_DIR, 'accounts')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
TRADING_DAYS_PER_YEAR = 365  # crypto markets trade every day

# ═══════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════
def find_default_data_file():
    """Find the most recent CSV in data/ directory."""
    pattern = os.path.join(DATA_DIR, '*.csv')
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    return files[0] if files else None


def load_market_data(filepath):
    """Load and validate the market data CSV."""
    print(f"\nLoading market data from: {filepath}")
    df = pd.read_csv(filepath)
    df.columns = df.columns.str.strip()

    required = ['DATE', 'SYMBOL', 'PRICE_USD', 'CEX_TRADING_VOLUME_24H_USD', 'CIRCULATING_TOKENS']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Data file missing columns: {missing}")

    df['DATE'] = pd.to_datetime(df['DATE'], format='mixed')
    df['SYMBOL'] = df['SYMBOL'].str.upper().str.strip()
    df = df.sort_values(['SYMBOL', 'DATE']).reset_index(drop=True)

    symbols = sorted(df['SYMBOL'].unique())
    date_range = f"{df['DATE'].min().strftime('%Y-%m-%d')} to {df['DATE'].max().strftime('%Y-%m-%d')}"
    print(f"  Tokens found: {', '.join(symbols)}")
    print(f"  Date range: {date_range}")
    print(f"  Total rows: {len(df):,}")
    return df


# ═══════════════════════════════════════════════════════════════════
# PER-TOKEN ANALYTICS (computed once, shared across accounts)
# ═══════════════════════════════════════════════════════════════════
def compute_token_analytics(df):
    """
    Compute per-token metrics from market data:
      - annualized_volatility: std(log returns) * sqrt(365)
      - avg_volume_30d: mean volume over 30 most recent days
      - avg_circulating_30d: mean circulating tokens over 30 most recent days
      - latest_price: most recent closing price
      - latest_date: date of most recent price
      - vwap_7d: 7-day volume-weighted average price
    Also returns the full returns DataFrame and annualized covariance matrix.
    """
    analytics = {}
    price_pivot = {}

    for symbol, gdf in df.groupby('SYMBOL'):
        gdf = gdf.sort_values('DATE').reset_index(drop=True)

        # Latest price
        latest_price = gdf.iloc[-1]['PRICE_USD']
        latest_date = gdf.iloc[-1]['DATE'].strftime('%Y-%m-%d')

        # 7-day VWAP
        recent_7 = gdf.tail(7)
        vwap_7d = (
            (recent_7['PRICE_USD'] * recent_7['CEX_TRADING_VOLUME_24H_USD']).sum()
            / recent_7['CEX_TRADING_VOLUME_24H_USD'].sum()
        )

        # 30-day averages
        recent_30 = gdf.tail(30)
        avg_volume_30d = recent_30['CEX_TRADING_VOLUME_24H_USD'].mean()
        avg_circulating_30d = recent_30['CIRCULATING_TOKENS'].mean()

        # 30-day average market cap
        avg_market_cap_30d = (recent_30['PRICE_USD'] * recent_30['CIRCULATING_TOKENS']).mean()

        # Annualized volatility from log returns
        prices = gdf['PRICE_USD'].values
        log_returns = np.diff(np.log(prices))
        annualized_vol = np.std(log_returns, ddof=1) * np.sqrt(TRADING_DAYS_PER_YEAR)

        # Store price series for covariance
        price_pivot[symbol] = gdf.set_index('DATE')['PRICE_USD']

        analytics[symbol] = {
            'latest_price': latest_price,
            'latest_date': latest_date,
            'vwap_7d': vwap_7d,
            'avg_volume_30d': avg_volume_30d,
            'avg_circulating_30d': avg_circulating_30d,
            'avg_market_cap_30d': avg_market_cap_30d,
            'annualized_volatility': annualized_vol,
        }

    # Build returns DataFrame for covariance matrix
    price_df = pd.DataFrame(price_pivot).dropna()
    returns_df = price_df.pct_change().dropna()
    cov_matrix = returns_df.cov() * TRADING_DAYS_PER_YEAR

    print("\n── Token Analytics ──")
    print(f"  {'Token':<8} {'Price':>12} {'Ann.Vol':>10} {'30d AvgVol':>18} {'30d Mkt Cap':>18}")
    for sym in sorted(analytics.keys()):
        a = analytics[sym]
        print(f"  {sym:<8} ${a['latest_price']:>11,.2f} {a['annualized_volatility']:>9.2%} "
              f"${a['avg_volume_30d']:>16,.0f} ${a['avg_market_cap_30d']:>16,.0f}")

    return analytics, returns_df, cov_matrix


# ═══════════════════════════════════════════════════════════════════
# ACCOUNT LOADING
# ═══════════════════════════════════════════════════════════════════
def load_account(filepath):
    """Load and validate an account JSON file."""
    with open(filepath, 'r') as f:
        raw = f.read()

    import re
    # Strip BOM and carriage returns
    raw = raw.strip().lstrip('\ufeff')
    raw = raw.replace('\r\n', '\n').replace('\r', '\n')
    # Fix trailing commas before closing braces
    raw = re.sub(r',\s*}', '}', raw)
    # Fix missing commas between key-value pairs (line ending with number/string, next line has quote)
    raw = re.sub(r'(\d)\s*\n(\s*")', r'\1,\n\2', raw)
    raw = re.sub(r'"\s*\n(\s*")', r'",\n\1', raw)
    # Fix unquoted string values (e.g., "account_name": deep_corp)
    raw = re.sub(r':\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*([,}\n])', r': "\1"\2', raw)

    acct = json.loads(raw)

    # Normalize
    acct['account_name'] = str(acct.get('account_name', os.path.basename(filepath).replace('.json', '')))
    acct['benchmark'] = acct.get('benchmark', 'custom').lower().strip().replace(' ', '_')
    acct['contribution'] = float(acct.get('contribution', 0))
    acct['pricing_method'] = acct.get('pricing_method', 'latest_close').lower().strip()

    # Normalize holdings keys to uppercase
    raw_holdings = acct.get('holdings', {})
    acct['holdings'] = {k.upper(): float(v) for k, v in raw_holdings.items()}

    # Normalize custom_weights if present
    if 'custom_weights' in acct:
        raw_cw = acct['custom_weights']
        acct['custom_weights'] = {k.upper(): float(v) for k, v in raw_cw.items()}

    return acct


def load_all_accounts():
    """Load all JSON account files from accounts/ directory."""
    pattern = os.path.join(ACCOUNTS_DIR, '*.json')
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"  No account files found in {ACCOUNTS_DIR}/")
        return []

    accounts = []
    for fp in files:
        try:
            acct = load_account(fp)
            accounts.append(acct)
            print(f"  Loaded: {acct['account_name']} ({acct['benchmark']}, "
                  f"{len(acct['holdings'])} assets)")
        except Exception as e:
            print(f"  WARNING: Skipping {fp}: {e}")
    return accounts


# ═══════════════════════════════════════════════════════════════════
# BENCHMARK WEIGHT CALCULATIONS
# ═══════════════════════════════════════════════════════════════════
def _normalize_weights(w):
    """Ensure weights sum to exactly 1.0 by adjusting the largest."""
    total = sum(w.values())
    if abs(total - 1.0) > 1e-8 and total > 0:
        largest = max(w, key=w.get)
        w[largest] = round(w[largest] + (1.0 - total), 10)
    return w


def calc_cap_weighted(tokens, analytics):
    """Market-cap weighted using 30-day average market cap."""
    caps = {}
    for t in tokens:
        if t == 'USDC' or t not in analytics:
            continue
        caps[t] = analytics[t]['avg_market_cap_30d']

    if not caps:
        raise ValueError("No market cap data for cap-weighted benchmark")

    total = sum(caps.values())
    weights = {t: round(v / total, 6) for t, v in caps.items()}
    weights['USDC'] = 0.0
    return _normalize_weights(weights)


def calc_volume_weighted(tokens, analytics):
    """Volume weighted using 30-day average trading volume."""
    vols = {}
    for t in tokens:
        if t == 'USDC' or t not in analytics:
            continue
        vols[t] = analytics[t]['avg_volume_30d']

    if not vols:
        raise ValueError("No volume data for volume-weighted benchmark")

    total = sum(vols.values())
    weights = {t: round(v / total, 6) for t, v in vols.items()}
    weights['USDC'] = 0.0
    return _normalize_weights(weights)


def calc_risk_parity(tokens, returns_df, cov_full):
    """Equal Risk Contribution (risk parity) using covariance matrix."""
    active = [t for t in tokens if t != 'USDC' and t in cov_full.columns]
    if len(active) < 2:
        raise ValueError("Risk parity requires at least 2 non-USDC assets with price data")

    cov = cov_full.loc[active, active].values
    n = len(active)

    def objective(w):
        port_var = w @ cov @ w
        port_std = np.sqrt(port_var)
        mrc = cov @ w / port_std
        rc = w * mrc
        target = rc.mean()
        return np.sum((rc - target) ** 2)

    x0 = np.ones(n) / n
    bounds = [(0.0, 1.0)] * n
    constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1.0}
    res = minimize(objective, x0, method='SLSQP', bounds=bounds, constraints=constraints,
                   options={'maxiter': 1000, 'ftol': 1e-12})

    if not res.success:
        print(f"    Warning: Risk-parity optimization did not fully converge: {res.message}")

    weights = {t: round(w, 6) for t, w in zip(active, res.x)}
    weights['USDC'] = 0.0
    return _normalize_weights(weights)


def calc_custom_weights(acct):
    """Use custom weights from account JSON."""
    cw = acct.get('custom_weights', {})
    if not cw:
        raise ValueError("benchmark is 'custom' but no custom_weights provided")
    total = sum(cw.values())
    if not (0.9999 <= total <= 1.0001):
        raise ValueError(f"custom_weights must sum to 1.0 (got {total:.6f})")
    return _normalize_weights(dict(cw))


def compute_benchmark_weights(acct, analytics, returns_df, cov_matrix):
    """Dispatch to the correct benchmark calculation."""
    benchmark = acct['benchmark']
    tokens = list(acct['holdings'].keys())

    if benchmark in ('cap_weighted', 'market_cap'):
        weights = calc_cap_weighted(tokens, analytics)
        desc = "Market-cap weighted (30-day average)"
    elif benchmark in ('volume_weighted', 'volume'):
        weights = calc_volume_weighted(tokens, analytics)
        desc = "Volume weighted (30-day average)"
    elif benchmark == 'risk_parity':
        weights = calc_risk_parity(tokens, returns_df, cov_matrix)
        desc = "Risk parity / Equal Risk Contribution (180-day covariance)"
    elif benchmark == 'custom':
        weights = calc_custom_weights(acct)
        desc = "Custom weights (from account config)"
    else:
        raise ValueError(f"Unknown benchmark: {benchmark}")

    # Ensure every held asset has a weight entry (default 0)
    for t in tokens:
        weights.setdefault(t, 0.0)

    return weights, desc


# ═══════════════════════════════════════════════════════════════════
# PRICING
# ═══════════════════════════════════════════════════════════════════
def get_price(token, method, analytics):
    """Get price for a token based on pricing method."""
    t = token.upper()
    if t == 'USDC':
        return 1.0, 'fixed'
    if t not in analytics:
        print(f"    Warning: No data for {t}, pricing at $0")
        return 0.0, 'no data'

    a = analytics[t]
    if method == 'vwap_7d':
        return a['vwap_7d'], f"7d VWAP to {a['latest_date']}"
    else:
        return a['latest_price'], a['latest_date']


# ═══════════════════════════════════════════════════════════════════
# REBALANCING CALCULATION
# ═══════════════════════════════════════════════════════════════════
def rebalance_account(acct, analytics, returns_df, cov_matrix):
    """Compute the full rebalancing for one account."""
    name = acct['account_name']
    method = acct['pricing_method']
    contribution = acct['contribution']

    # Price each holding
    prices = {}
    for token in acct['holdings']:
        price, date_str = get_price(token, method, analytics)
        prices[token] = {'price': price, 'date': date_str}

    # Current portfolio
    rows = []
    for token, units in acct['holdings'].items():
        p = prices[token]['price']
        value = units * p
        rows.append({'asset': token, 'units': units, 'market_price': p, 'value': value})

    total_value = sum(r['value'] for r in rows)
    adjusted_total = total_value + contribution

    # Benchmark weights
    weights, bench_desc = compute_benchmark_weights(acct, analytics, returns_df, cov_matrix)

    # Rebalancing trades
    results = []
    for r in rows:
        asset = r['asset']
        current_value = r['value']
        current_weight = current_value / total_value if total_value > 0 else 0
        model_weight = weights.get(asset, 0.0)
        target_value = model_weight * adjusted_total
        trade_value = target_value - current_value
        swap_units = trade_value / r['market_price'] if r['market_price'] > 0 else 0

        results.append({
            'asset': asset,
            'units': r['units'],
            'market_price': r['market_price'],
            'value': current_value,
            'portfolio_weight': current_weight,
            'model_weight': model_weight,
            'difference': current_weight - model_weight,
            'trade_value': trade_value,
            'swap_units': swap_units,
        })

    # Validate trades sum to contribution
    total_trade = sum(r['trade_value'] for r in results)
    tolerance = max(1e-6 * adjusted_total, 0.01)
    if abs(total_trade - contribution) > tolerance:
        print(f"    WARNING: Trade sum ({total_trade:,.2f}) != contribution ({contribution:,.2f})")

    return {
        'account_name': name,
        'benchmark_desc': bench_desc,
        'pricing_method': method,
        'contribution': contribution,
        'total_value': total_value,
        'adjusted_total': adjusted_total,
        'total_trade': total_trade,
        'results': sorted(results, key=lambda x: x['asset']),
        'weights': weights,
        'prices': prices,
    }


# ═══════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════
def write_excel(rebal, analytics, output_path):
    """Generate the formatted Excel rebalancing report."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rebalance'

    # Styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    blue_font = Font(name='Arial', size=10, color='0000FF')
    meta_font = Font(name='Arial', size=9, italic=True, color='555555')
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    light_fill = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Column widths
    col_widths = {'A': 14, 'B': 16, 'C': 16, 'D': 18, 'E': 16, 'F': 16, 'G': 14, 'H': 18, 'I': 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: Title
    ws['A1'] = 'Portfolio Rebalancing Model'
    ws['A1'].font = title_font

    # ── Row 3: Account name
    ws['A3'] = f"Account: {rebal['account_name']}"
    ws['A3'].font = Font(name='Arial', size=11, bold=True)

    # ── Row 4: Header groupings
    groups = [('A4', 'eligible'), ('B4', 'holdings'), ('D4', 'percent'), ('G4', 'recommended trade')]
    for cell, val in groups:
        ws[cell] = val
        ws[cell].font = header_font
        ws[cell].fill = header_fill

    # ── Row 5: Column headers
    headers = ['assets', 'units', 'market price', '$ value', 'portfolio weight',
               'model weight', 'difference', 'trade value', 'swap units']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # ── Data rows (starting row 7)
    row = 7
    results = rebal['results']
    for r in results:
        ws.cell(row=row, column=1, value=r['asset']).font = normal_font
        ws.cell(row=row, column=2, value=r['units']).font = blue_font
        ws.cell(row=row, column=2).number_format = '#,##0.000000'
        ws.cell(row=row, column=3, value=r['market_price']).font = normal_font
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=r['value']).font = normal_font
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=r['portfolio_weight']).font = normal_font
        ws.cell(row=row, column=5).number_format = '0.0000%'
        ws.cell(row=row, column=6, value=r['model_weight']).font = normal_font
        ws.cell(row=row, column=6).number_format = '0.0000%'
        ws.cell(row=row, column=7, value=r['difference']).font = normal_font
        ws.cell(row=row, column=7).number_format = '0.0000%'
        ws.cell(row=row, column=8, value=r['trade_value']).font = normal_font
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        ws.cell(row=row, column=9, value=r['swap_units']).font = normal_font
        ws.cell(row=row, column=9).number_format = '#,##0.000000'

        # Alternate row shading
        if (row - 7) % 2 == 1:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = light_fill
        row += 1

    # ── Contribution row
    row += 1
    ws.cell(row=row, column=1, value='contribution').font = Font(name='Arial', size=10, italic=True)
    ws.cell(row=row, column=4, value=rebal['contribution']).font = normal_font
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    if rebal['total_value'] > 0:
        ws.cell(row=row, column=5, value=rebal['contribution'] / rebal['total_value']).font = normal_font
        ws.cell(row=row, column=5).number_format = '0.0000%'
    row += 1

    # ── Total row
    row += 1
    total_border = Border(top=Side(style='double'), bottom=Side(style='double'))
    ws.cell(row=row, column=1, value='total').font = header_font
    ws.cell(row=row, column=4, value=rebal['adjusted_total']).font = header_font
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=5, value=1.0).font = header_font
    ws.cell(row=row, column=5).number_format = '0.0000%'
    ws.cell(row=row, column=6, value=1.0).font = header_font
    ws.cell(row=row, column=6).number_format = '0.0000%'
    ws.cell(row=row, column=7, value=0.0).font = header_font
    ws.cell(row=row, column=7).number_format = '0.0000%'
    ws.cell(row=row, column=8, value=rebal['total_trade']).font = header_font
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = total_border

    # ── Metadata section
    row += 3
    ws.cell(row=row, column=1, value='Benchmark:').font = header_font
    row += 1
    ws.cell(row=row, column=1, value=rebal['benchmark_desc']).font = meta_font
    row += 2

    ws.cell(row=row, column=1, value='Pricing Method:').font = header_font
    row += 1
    ws.cell(row=row, column=1, value=rebal['pricing_method']).font = meta_font
    row += 2

    ws.cell(row=row, column=1, value='Prices Used:').font = header_font
    row += 1
    for token in sorted(rebal['prices'].keys()):
        info = rebal['prices'][token]
        ws.cell(row=row, column=1,
                value=f"{token}: ${info['price']:,.2f} ({info['date']})").font = meta_font
        row += 1

    # ── Token analytics section
    row += 2
    ws.cell(row=row, column=1, value='Token Analytics:').font = header_font
    row += 1
    analytic_headers = ['Token', 'Ann. Volatility', '30d Avg Volume', '30d Avg Circ. Supply']
    for i, h in enumerate(analytic_headers, 1):
        ws.cell(row=row, column=i, value=h).font = header_font
        ws.cell(row=row, column=i).fill = header_fill
    row += 1

    for token in sorted(rebal['prices'].keys()):
        if token == 'USDC' or token not in analytics:
            continue
        a = analytics[token]
        ws.cell(row=row, column=1, value=token).font = normal_font
        ws.cell(row=row, column=2, value=a['annualized_volatility']).font = normal_font
        ws.cell(row=row, column=2).number_format = '0.00%'
        ws.cell(row=row, column=3, value=a['avg_volume_30d']).font = normal_font
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=a['avg_circulating_30d']).font = normal_font
        ws.cell(row=row, column=4).number_format = '#,##0'
        row += 1

    # ── Model weights detail
    row += 2
    ws.cell(row=row, column=1, value='Model Weights:').font = header_font
    row += 1
    for token in sorted(rebal['weights'].keys()):
        w = rebal['weights'][token]
        if w > 0:
            ws.cell(row=row, column=1, value=token).font = normal_font
            ws.cell(row=row, column=2, value=w).font = normal_font
            ws.cell(row=row, column=2).number_format = '0.0000%'
            row += 1

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Multi-Account Crypto Portfolio Rebalancer")
    print("=" * 60)

    # Ensure directories exist
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(ACCOUNTS_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Prompt for data file
    default_file = find_default_data_file()
    if default_file:
        prompt_str = f"Enter path to market data CSV [{default_file}]: "
    else:
        prompt_str = "Enter path to market data CSV: "

    user_input = input(prompt_str).strip()
    data_path = user_input if user_input else default_file
    if not data_path or not os.path.exists(data_path):
        print(f"ERROR: Data file not found: {data_path}")
        sys.exit(1)

    # Load market data
    df = load_market_data(data_path)

    # Compute analytics
    analytics, returns_df, cov_matrix = compute_token_analytics(df)

    # Load accounts
    print(f"\nLoading accounts from: {ACCOUNTS_DIR}/")
    accounts = load_all_accounts()
    if not accounts:
        print("No accounts to process. Exiting.")
        sys.exit(0)

    # Process each account
    today_str = datetime.now().strftime('%Y%m%d')
    output_files = []

    print("\n" + "=" * 60)
    print("  Processing Accounts")
    print("=" * 60)

    for acct in accounts:
        name = acct['account_name']
        print(f"\n── Account: {name} ({acct['benchmark']}) ──")
        try:
            rebal = rebalance_account(acct, analytics, returns_df, cov_matrix)

            output_path = os.path.join(OUTPUT_DIR, f"rebal_{name}_{today_str}.xlsx")
            write_excel(rebal, analytics, output_path)
            output_files.append(output_path)

            print(f"    Portfolio value: ${rebal['total_value']:,.2f}")
            print(f"    Contribution:   ${rebal['contribution']:,.2f}")
            print(f"    Adjusted total: ${rebal['adjusted_total']:,.2f}")
            print(f"    Benchmark:      {rebal['benchmark_desc']}")
            print(f"    Output:         {output_path}")

            # Print weight summary
            print(f"    Weights: ", end='')
            wt_parts = [f"{t}={w:.2%}" for t, w in sorted(rebal['weights'].items()) if w > 0]
            print(', '.join(wt_parts))

        except Exception as e:
            print(f"    ERROR processing {name}: {e}")
            import traceback
            traceback.print_exc()

    # Summary
    print("\n" + "=" * 60)
    print(f"  Complete! Generated {len(output_files)} rebalancing report(s)")
    for f in output_files:
        print(f"    → {f}")
    print("=" * 60)


if __name__ == '__main__':
    main()